import frappe
import io
from frappe.utils import flt, cint, now_datetime
from vcl_sales_dashboard.api.collections_utils import (
    normalise_header, map_headers, validate_required_columns,
    match_customer, resolve_sales_rep_assignment, safe_flt,
    CURRENCY_FIELDS, get_rep_label_map, resolve_rep_user_from_label,
    resolve_sales_rep_user,
)
import json


# ── Helpers ──────────────────────────────────────────────────────────

def extract_ageing_dataframe(file_content):
    """Parse Excel bytes into rows from the Ageing Report sheet.
    Returns (data_rows, raw_headers, sheet_name)."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)

    sheet_name = None
    for name in wb.sheetnames:
        if name.strip().lower() == "ageing report":
            sheet_name = name
            break

    if not sheet_name:
        frappe.throw(
            f"Sheet 'Ageing Report' not found. Available sheets: {', '.join(wb.sheetnames)}"
        )

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        frappe.throw("The 'Ageing Report' sheet is empty.")

    headers = [str(h).strip() if h else "" for h in rows[0]]
    data = []
    for row in rows[1:]:
        row_dict = {}
        for i, val in enumerate(row):
            if i < len(headers) and headers[i]:
                row_dict[headers[i]] = val
        data.append(row_dict)

    return data, headers, sheet_name


def check_duplicate_period(period_end, replace_flag, exclude_name=None):
    """Check if a submission already exists for this period end."""
    conditions = {
        "period_end": period_end,
        "status": ["in", ["Validated", "Processed"]],
    }
    if exclude_name:
        conditions["name"] = ["!=", exclude_name]

    existing = frappe.db.get_all("Collections Submission", filters=conditions, pluck="name")

    if existing and not cint(replace_flag):
        frappe.throw(
            f"A submission already exists for period ending {period_end}: "
            f"{', '.join(existing)}. Check 'Replace Existing Month' to overwrite."
        )
    return existing


def delete_existing_snapshots(submission_names):
    """Delete snapshot rows for given submission names."""
    for sub_name in submission_names:
        frappe.db.sql("""
            DELETE fu FROM `tabCollections Follow Up` fu
            INNER JOIN `tabCollections Customer Snapshot` cs
                ON fu.collections_customer_snapshot = cs.name
            WHERE cs.collections_submission = %(sub)s
        """, {"sub": sub_name})

        frappe.db.sql("""
            DELETE FROM `tabCollections Customer Snapshot`
            WHERE collections_submission = %(sub)s
        """, {"sub": sub_name})

        frappe.db.set_value("Collections Submission", sub_name, "status", "Draft")

    frappe.db.commit()


def build_snapshot_doc(submission_doc, row_dict, header_map, log_lines, label_map=None):
    """Create one Collections Customer Snapshot from a mapped row."""
    mapped = {}
    for excel_col, field in header_map.items():
        mapped[field] = row_dict.get(excel_col)

    excel_name = (mapped.get("excel_customer_name") or "").strip()
    if not excel_name:
        return None, "Skipped: empty customer name"

    customer, match_status = match_customer(excel_name)
    rep_info = resolve_sales_rep_assignment(customer, submission_doc.period_end)

    excel_rep_label = (mapped.get("excel_sales_representative") or "").strip() if mapped.get("excel_sales_representative") else ""

    # Fallback: if CSR assignment didn't resolve a user, try Excel label
    if not rep_info["sales_rep_user"] and excel_rep_label:
        resolved = resolve_sales_rep_user(excel_rep_label)
        if resolved:
            rep_info["sales_rep_user"] = resolved
            if not rep_info["assigned_sales_representative"]:
                rep_info["assigned_sales_representative"] = excel_rep_label
            if rep_info["assignment_match_status"] in ("Imported With Warning", "No Valid Assignment For Period"):
                rep_info["assignment_match_status"] = "Matched"

    # Final safety: sales_rep_user must be a real User or blank
    if rep_info["sales_rep_user"] and not frappe.db.exists("User", rep_info["sales_rep_user"]):
        rep_info["sales_rep_user"] = None

    # If user is still unresolved, force warning status
    if rep_info["assigned_sales_representative"] and not rep_info["sales_rep_user"]:
        rep_info["assignment_match_status"] = "Imported With Warning"

    snapshot = frappe.new_doc("Collections Customer Snapshot")
    snapshot.collections_submission = submission_doc.name
    snapshot.period_end = submission_doc.period_end
    snapshot.snapshot_label = submission_doc.submission_label

    snapshot.excel_customer_name = excel_name
    snapshot.customer = customer
    snapshot.customer_name_display = customer or excel_name
    snapshot.customer_match_status = match_status

    snapshot.excel_sales_representative = excel_rep_label
    snapshot.assigned_sales_representative = rep_info["assigned_sales_representative"] or excel_rep_label or None
    snapshot.sales_rep_user = rep_info["sales_rep_user"] or None
    snapshot.customer_sales_rep_assignment = rep_info["customer_sales_rep_assignment"]
    snapshot.assignment_match_status = rep_info["assignment_match_status"]

    snapshot.terms = (mapped.get("terms") or "").strip() if mapped.get("terms") else ""

    for field in CURRENCY_FIELDS:
        setattr(snapshot, field, safe_flt(mapped.get(field)))

    snapshot.latest_follow_up_status = "Not Contacted"
    snapshot.insert(ignore_permissions=True)

    warnings = []
    if match_status != "Matched":
        warnings.append(f"customer: {match_status}")
    if rep_info["assignment_match_status"] != "Matched":
        warnings.append(f"rep: {rep_info['assignment_match_status']}")
    if excel_rep_label and not snapshot.sales_rep_user:
        warnings.append(f"rep user unresolved: '{excel_rep_label}'")
    elif snapshot.assigned_sales_representative and not snapshot.sales_rep_user:
        warnings.append(f"rep user unresolved from assignment: '{snapshot.assigned_sales_representative}'")

    if warnings:
        log_lines.append(f"  [{excel_name}] {', '.join(warnings)}")

    return snapshot, None


def update_submission_summary(submission_doc):
    """Recalculate totals and counts on the submission doc."""
    snapshots = frappe.get_all(
        "Collections Customer Snapshot",
        filters={"collections_submission": submission_doc.name},
        fields=[
            "customer_match_status", "assignment_match_status",
            "total_balance", "overdue_amount", "overdue_30_amount",
            "due_current_month", "due_next_month", "pd_cheques_cm",
        ],
    )

    submission_doc.rows_imported = len(snapshots)
    submission_doc.customers_matched = sum(1 for s in snapshots if s.customer_match_status == "Matched")
    submission_doc.customers_unmatched = sum(1 for s in snapshots if s.customer_match_status != "Matched")
    submission_doc.assignment_matched = sum(1 for s in snapshots if s.assignment_match_status == "Matched")
    submission_doc.assignment_unmatched = sum(1 for s in snapshots if s.assignment_match_status == "No Valid Assignment For Period")
    submission_doc.multiple_assignments_found = sum(1 for s in snapshots if s.assignment_match_status == "Multiple Assignments Found")

    submission_doc.total_balance = sum(flt(s.total_balance) for s in snapshots)
    submission_doc.total_overdue = sum(flt(s.overdue_amount) for s in snapshots)
    submission_doc.total_overdue_30 = sum(flt(s.overdue_30_amount) for s in snapshots)
    submission_doc.total_due_current_month = sum(flt(s.due_current_month) for s in snapshots)
    submission_doc.total_due_next_month = sum(flt(s.due_next_month) for s in snapshots)
    submission_doc.total_pd_cheques_cm = sum(flt(s.pd_cheques_cm) for s in snapshots)

    submission_doc.save(ignore_permissions=True)


# ── Whitelisted API ──────────────────────────────────────────────────

@frappe.whitelist()
def validate_submission_file():
    """Validate an uploaded Excel file without creating any records."""
    try:
        file = frappe.request.files.get("file")
        period_end = frappe.form_dict.get("period_end")

        if not file:
            return {"status": "error", "message": "No file uploaded."}
        if not period_end:
            return {"status": "error", "message": "Period End is required."}

        file_content = file.read()
        filename = file.filename or ""

        if not filename.lower().endswith((".xlsx", ".xls")):
            return {"status": "error", "message": "File must be an Excel file (.xlsx or .xls)."}

        data, raw_headers, sheet_name = extract_ageing_dataframe(file_content)
        header_map = map_headers(raw_headers)

        missing = validate_required_columns(header_map)
        if missing:
            return {"status": "error", "message": f"Missing required columns: {', '.join(missing)}"}

        customer_col = None
        for raw_col, field in header_map.items():
            if field == "excel_customer_name":
                customer_col = raw_col
                break

        valid_rows = [r for r in data if r.get(customer_col, "")]
        if not valid_rows:
            return {"status": "error", "message": "No customer rows found in the Ageing Report sheet."}

        # Check for unresolved sales rep labels
        rep_col = None
        for raw_col, field in header_map.items():
            if field == "excel_sales_representative":
                rep_col = raw_col
                break

        unresolved_labels = []
        if rep_col:
            label_map = get_rep_label_map()
            seen_labels = set()
            for row in valid_rows:
                label = (str(row.get(rep_col, "") or "")).strip()
                if label and label not in seen_labels:
                    seen_labels.add(label)
                    # Check if this label is already a valid User email
                    if frappe.db.exists("User", label):
                        continue
                    # Check if it's in the mapping table
                    if label in label_map:
                        continue
                    unresolved_labels.append(label)

        if unresolved_labels:
            return {
                "status": "needs_mapping",
                "message": f"{len(unresolved_labels)} sales rep label(s) could not be resolved to ERPNext users.",
                "unresolved_sales_rep_labels": sorted(unresolved_labels),
                "rows_found": len(valid_rows),
                "columns_mapped": len(header_map),
                "sheet_name": sheet_name,
                "filename": filename,
            }

        return {
            "status": "ok",
            "message": "Validation successful.",
            "rows_found": len(valid_rows),
            "columns_mapped": len(header_map),
            "sheet_name": sheet_name,
            "filename": filename,
            "mapped_fields": sorted(header_map.values()),
        }

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Collections validation error")
        return {"status": "error", "message": str(e)}


@frappe.whitelist()
def process_collections_submission():
    """Create a submission header and process all rows from the uploaded file."""
    try:
        file = frappe.request.files.get("file")
        period_end = frappe.form_dict.get("period_end")
        replace_existing = cint(frappe.form_dict.get("replace_existing_month", 0))
        notes = frappe.form_dict.get("submission_notes", "")

        if not file:
            return {"status": "error", "message": "No file uploaded."}
        if not period_end:
            return {"status": "error", "message": "Period End is required."}

        file_content = file.read()
        filename = file.filename or ""

        if not filename.lower().endswith((".xlsx", ".xls")):
            return {"status": "error", "message": "File must be an Excel file (.xlsx or .xls)."}

        # Parse Excel
        data, raw_headers, sheet_name = extract_ageing_dataframe(file_content)
        header_map = map_headers(raw_headers)

        missing = validate_required_columns(header_map)
        if missing:
            return {"status": "error", "message": f"Missing required columns: {', '.join(missing)}"}

        # Check for duplicate period
        existing = check_duplicate_period(period_end, replace_existing)
        if existing:
            delete_existing_snapshots(existing)

        # Create submission header (submission_label is auto-derived in validate)
        sub = frappe.new_doc("Collections Submission")
        sub.period_end = period_end
        sub.status = "Validated"
        sub.submitted_by = frappe.session.user
        sub.submitted_on = now_datetime()
        sub.source_filename = filename
        sub.submission_notes = notes
        sub.replace_existing_month = replace_existing
        sub.rows_found = len(data)
        sub.insert(ignore_permissions=True)
        frappe.db.commit()

        # Find customer column
        customer_col = None
        for raw_col, field in header_map.items():
            if field == "excel_customer_name":
                customer_col = raw_col
                break

        # Load rep label mappings for fallback resolution
        label_map = get_rep_label_map()

        # Process rows
        log_lines = []
        imported = 0
        skipped = 0

        for row in data:
            cust_name = (str(row.get(customer_col, "") or "")).strip()
            if not cust_name:
                skipped += 1
                continue

            snapshot, err = build_snapshot_doc(sub, row, header_map, log_lines, label_map)
            if snapshot:
                imported += 1
            else:
                skipped += 1
                if err:
                    log_lines.append(f"  [row] {err}")

        # Update submission
        sub.reload()
        sub.rows_found = len(data)
        sub.rows_imported = imported
        sub.rows_skipped = skipped
        sub.status = "Processed"
        sub.validation_log = (
            f"Processing complete.\n"
            f"Sheet: {sheet_name}\n"
            f"Rows processed: {imported}\n"
            f"Rows skipped: {skipped}\n"
        )
        if log_lines:
            sub.validation_log += "\nWarnings:\n" + "\n".join(log_lines)

        sub.save(ignore_permissions=True)
        update_submission_summary(sub)
        frappe.db.commit()

        return {
            "status": "ok",
            "message": f"Submission complete. {imported} rows processed, {skipped} skipped.",
            "submission_name": sub.name,
            "rows_imported": imported,
            "rows_skipped": skipped,
            "warnings": len(log_lines),
        }

    except Exception as e:
        frappe.db.rollback()
        frappe.log_error(frappe.get_traceback(), "Collections submission error")
        return {"status": "error", "message": str(e)}


@frappe.whitelist()
def rebuild_submission_totals(submission_name):
    """Recalculate the totals on a submission from its snapshot rows."""
    try:
        doc = frappe.get_doc("Collections Submission", submission_name)
        update_submission_summary(doc)
        return {"status": "ok", "message": "Totals recalculated."}
    except Exception as e:
        return {"status": "error", "message": str(e)}


@frappe.whitelist()
def get_submission_log(submission_name):
    """Return the validation log for a submission."""
    try:
        doc = frappe.get_doc("Collections Submission", submission_name)
        return {
            "status": "ok",
            "data": {
                "name": doc.name,
                "status": doc.status,
                "validation_log": doc.validation_log,
                "rows_found": doc.rows_found,
                "rows_imported": doc.rows_imported,
                "rows_skipped": doc.rows_skipped,
                "customers_matched": doc.customers_matched,
                "customers_unmatched": doc.customers_unmatched,
                "assignment_matched": doc.assignment_matched,
                "assignment_unmatched": doc.assignment_unmatched,
                "multiple_assignments_found": doc.multiple_assignments_found,
            },
        }
    except Exception as e:
        return {"status": "error", "message": str(e)}


@frappe.whitelist()
def save_rep_mappings(mappings=None):
    """Save sales rep label → ERPNext User mappings.
    Accepts a JSON array of {label, user} objects.
    Creates or updates Sales Rep User Mapping records."""
    try:
        if isinstance(mappings, str):
            mappings = json.loads(mappings)
        if not mappings:
            return {"status": "error", "message": "No mappings provided."}

        saved = 0
        for m in mappings:
            label = (m.get("label") or "").strip()
            user = (m.get("user") or "").strip()
            if not label or not user:
                continue

            if not frappe.db.exists("User", user):
                continue

            if frappe.db.exists("Sales Rep User Mapping", label):
                frappe.db.set_value("Sales Rep User Mapping", label, "user", user)
            else:
                doc = frappe.new_doc("Sales Rep User Mapping")
                doc.sales_rep_label = label
                doc.user = user
                doc.insert(ignore_permissions=True)
            saved += 1

        frappe.db.commit()
        return {"status": "ok", "message": f"{saved} mapping(s) saved.", "saved": saved}

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Save rep mappings error")
        return {"status": "error", "message": str(e)}


@frappe.whitelist()
def get_user_list():
    """Return list of active users for the mapping dropdown."""
    try:
        users = frappe.get_all(
            "User",
            filters={"enabled": 1, "user_type": "System User"},
            fields=["name", "full_name"],
            order_by="full_name asc",
            limit_page_length=200,
            ignore_permissions=True,
        )
        return {"status": "ok", "data": users}
    except Exception as e:
        return {"status": "error", "message": str(e)}


@frappe.whitelist()
def get_past_submissions():
    """Return past submissions list. Uses ignore_permissions since the
    page controller already enforces role access."""
    try:
        result = frappe.get_all(
            "Collections Submission",
            filters={"status": ["in", ["Validated", "Processed", "Failed", "Draft"]]},
            fields=[
                "name", "submission_label", "period_end", "status",
                "rows_imported", "total_balance", "total_overdue", "submitted_on",
            ],
            order_by="period_end desc",
            limit_page_length=20,
            ignore_permissions=True,
        )
        return {"status": "ok", "data": result}
    except Exception as e:
        return {"status": "error", "message": str(e)}
