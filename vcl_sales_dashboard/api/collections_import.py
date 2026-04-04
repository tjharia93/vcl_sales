import frappe
import json
from frappe.utils import flt, cint, now_datetime
from vcl_sales_dashboard.api.collections_utils import (
    get_file_path, normalise_header, map_headers, validate_required_columns,
    match_customer, resolve_sales_rep_assignment, safe_flt,
    CURRENCY_FIELDS, COLUMN_MAP,
)


# ── Helpers ──────────────────────────────────────────────────────────

def get_import_doc(import_name):
    """Fetch and return the Collections Import doc."""
    return frappe.get_doc("Collections Import", import_name)


def load_workbook_dataframe(file_path):
    """Load the Ageing Report sheet from the Excel workbook as a list of dicts."""
    import openpyxl

    wb = openpyxl.load_workbook(file_path, data_only=True)

    # Try exact name first, then case-insensitive
    sheet_name = None
    for name in wb.sheetnames:
        if name == "Ageing Report":
            sheet_name = name
            break
    if not sheet_name:
        for name in wb.sheetnames:
            if name.strip().lower() == "ageing report":
                sheet_name = name
                break

    if not sheet_name:
        frappe.throw(
            f"Sheet 'Ageing Report' not found. Available sheets: {', '.join(wb.sheetnames)}"
        )

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        frappe.throw("The 'Ageing Report' sheet is empty.")

    headers = [str(h).strip() if h else "" for h in rows[0]]
    data = []
    for row in rows[1:]:
        row_dict = {}
        for i, val in enumerate(row):
            if i < len(headers) and headers[i]:
                row_dict[headers[i]] = val
        data.append(row_dict)

    return data, headers, sheet_name


def check_duplicate_period(period_start, period_end, replace_flag, exclude_name=None):
    """Check if an import already exists for this period. Raises if blocked."""
    conditions = {
        "period_start": period_start,
        "period_end": period_end,
        "status": ["in", ["Validated", "Imported"]],
    }
    if exclude_name:
        conditions["name"] = ["!=", exclude_name]

    existing = frappe.db.get_all("Collections Import", filters=conditions, pluck="name")

    if existing and not cint(replace_flag):
        frappe.throw(
            f"An import already exists for period {period_start} to {period_end}: "
            f"{', '.join(existing)}. Check 'Replace Existing Month' to overwrite."
        )
    return existing


def delete_existing_snapshots(import_names):
    """Delete snapshot rows for given import names."""
    for imp_name in import_names:
        # Delete follow-ups first
        frappe.db.sql("""
            DELETE fu FROM `tabCollections Follow Up` fu
            INNER JOIN `tabCollections Customer Snapshot` cs
                ON fu.collections_customer_snapshot = cs.name
            WHERE cs.collections_import = %(imp)s
        """, {"imp": imp_name})

        # Delete snapshots
        frappe.db.sql("""
            DELETE FROM `tabCollections Customer Snapshot`
            WHERE collections_import = %(imp)s
        """, {"imp": imp_name})

        # Mark old import as replaced
        frappe.db.set_value("Collections Import", imp_name, "status", "Draft")

    frappe.db.commit()


def create_snapshot_row(import_doc, row_dict, header_map, log_lines):
    """Create one Collections Customer Snapshot from a mapped Excel row."""
    # Extract mapped values
    mapped = {}
    for excel_col, field in header_map.items():
        mapped[field] = row_dict.get(excel_col)

    excel_name = (mapped.get("excel_customer_name") or "").strip()
    if not excel_name:
        return None, "Skipped: empty customer name"

    # Match customer
    customer, match_status = match_customer(excel_name)

    # Resolve rep assignment
    rep_info = resolve_sales_rep_assignment(
        customer, import_doc.period_start, import_doc.period_end
    )

    # Build snapshot doc
    snapshot = frappe.new_doc("Collections Customer Snapshot")
    snapshot.collections_import = import_doc.name
    snapshot.period_start = import_doc.period_start
    snapshot.period_end = import_doc.period_end
    snapshot.snapshot_label = import_doc.import_label

    # Customer matching
    snapshot.excel_customer_name = excel_name
    snapshot.customer = customer
    snapshot.customer_name_display = customer or excel_name
    snapshot.customer_match_status = match_status

    # Rep resolution
    snapshot.excel_sales_representative = (mapped.get("excel_sales_representative") or "").strip()
    snapshot.assigned_sales_representative = rep_info["assigned_sales_representative"]
    snapshot.sales_rep_user = rep_info["sales_rep_user"]
    snapshot.customer_sales_rep_assignment = rep_info["customer_sales_rep_assignment"]
    snapshot.assignment_match_status = rep_info["assignment_match_status"]

    # Terms
    snapshot.terms = (mapped.get("terms") or "").strip() if mapped.get("terms") else ""

    # Currency fields
    for field in CURRENCY_FIELDS:
        setattr(snapshot, field, safe_flt(mapped.get(field)))

    # Default dashboard state
    snapshot.latest_follow_up_status = "Not Contacted"

    snapshot.insert(ignore_permissions=True)

    # Build log line
    warnings = []
    if match_status != "Matched":
        warnings.append(f"customer: {match_status}")
    if rep_info["assignment_match_status"] not in ("Matched",):
        warnings.append(f"rep: {rep_info['assignment_match_status']}")

    if warnings:
        log_lines.append(f"  [{excel_name}] {', '.join(warnings)}")

    return snapshot, None


def update_import_summary(import_doc):
    """Recalculate totals and counts on the import doc from its snapshots."""
    snapshots = frappe.get_all(
        "Collections Customer Snapshot",
        filters={"collections_import": import_doc.name},
        fields=[
            "customer_match_status", "assignment_match_status",
            "total_balance", "overdue_amount", "overdue_30_amount",
            "due_current_month", "due_next_month", "pd_cheques_cm",
        ],
    )

    import_doc.rows_imported = len(snapshots)
    import_doc.customers_matched = sum(
        1 for s in snapshots if s.customer_match_status == "Matched"
    )
    import_doc.customers_unmatched = sum(
        1 for s in snapshots if s.customer_match_status != "Matched"
    )
    import_doc.assignment_matched = sum(
        1 for s in snapshots if s.assignment_match_status == "Matched"
    )
    import_doc.assignment_unmatched = sum(
        1 for s in snapshots if s.assignment_match_status == "No Valid Assignment For Period"
    )
    import_doc.multiple_assignments_found = sum(
        1 for s in snapshots if s.assignment_match_status == "Multiple Assignments Found"
    )

    import_doc.total_balance = sum(flt(s.total_balance) for s in snapshots)
    import_doc.total_overdue = sum(flt(s.overdue_amount) for s in snapshots)
    import_doc.total_overdue_30 = sum(flt(s.overdue_30_amount) for s in snapshots)
    import_doc.total_due_current_month = sum(flt(s.due_current_month) for s in snapshots)
    import_doc.total_due_next_month = sum(flt(s.due_next_month) for s in snapshots)
    import_doc.total_pd_cheques_cm = sum(flt(s.pd_cheques_cm) for s in snapshots)

    import_doc.save(ignore_permissions=True)


# ── Whitelisted API functions ────────────────────────────────────────

@frappe.whitelist()
def validate_collections_file(import_name):
    """Validate the uploaded Excel file without importing rows."""
    try:
        doc = get_import_doc(import_name)
        file_path = get_file_path(doc.source_file)

        data, raw_headers, sheet_name = load_workbook_dataframe(file_path)
        header_map = map_headers(raw_headers)

        # Check required columns
        missing = validate_required_columns(header_map)
        if missing:
            doc.status = "Failed"
            doc.validation_log = f"Missing required columns: {', '.join(missing)}"
            doc.save(ignore_permissions=True)
            return {"status": "error", "message": doc.validation_log}

        # Filter to rows that have a customer name
        customer_col = None
        for raw_col, field in header_map.items():
            if field == "excel_customer_name":
                customer_col = raw_col
                break

        valid_rows = [r for r in data if r.get(customer_col, "")]
        if not valid_rows:
            doc.status = "Failed"
            doc.validation_log = "No customer rows found in the Ageing Report sheet."
            doc.save(ignore_permissions=True)
            return {"status": "error", "message": doc.validation_log}

        # Check for duplicate period
        try:
            check_duplicate_period(
                doc.period_start, doc.period_end, doc.replace_existing_month, doc.name
            )
        except Exception as e:
            doc.status = "Failed"
            doc.validation_log = str(e)
            doc.save(ignore_permissions=True)
            return {"status": "error", "message": str(e)}

        # Update doc
        doc.sheet_name_used = sheet_name
        doc.rows_found = len(valid_rows)
        doc.status = "Validated"
        doc.validation_log = (
            f"Validated OK.\n"
            f"Sheet: {sheet_name}\n"
            f"Columns mapped: {len(header_map)}\n"
            f"Customer rows found: {len(valid_rows)}\n"
            f"Mapped fields: {', '.join(sorted(header_map.values()))}"
        )
        doc.save(ignore_permissions=True)

        return {
            "status": "ok",
            "message": "Validation successful.",
            "rows_found": len(valid_rows),
            "columns_mapped": len(header_map),
            "sheet_name": sheet_name,
        }

    except Exception as e:
        frappe.log_error(f"Collections validation error: {e}")
        try:
            doc = get_import_doc(import_name)
            doc.status = "Failed"
            doc.validation_log = str(e)
            doc.save(ignore_permissions=True)
        except Exception:
            pass
        return {"status": "error", "message": str(e)}


@frappe.whitelist()
def import_collections_snapshot(import_name):
    """Import all rows from the validated Excel file as snapshot records."""
    try:
        doc = get_import_doc(import_name)

        if doc.status not in ("Validated", "Draft"):
            return {"status": "error", "message": f"Import status is '{doc.status}', expected 'Validated'."}

        file_path = get_file_path(doc.source_file)
        data, raw_headers, sheet_name = load_workbook_dataframe(file_path)
        header_map = map_headers(raw_headers)

        # Check required
        missing = validate_required_columns(header_map)
        if missing:
            doc.status = "Failed"
            doc.validation_log = f"Missing required columns: {', '.join(missing)}"
            doc.save(ignore_permissions=True)
            return {"status": "error", "message": doc.validation_log}

        # Handle duplicate period / replacement
        existing = check_duplicate_period(
            doc.period_start, doc.period_end, doc.replace_existing_month, doc.name
        )
        if existing:
            delete_existing_snapshots(existing)

        # Delete any snapshots from a previous attempt on this same import
        frappe.db.sql("""
            DELETE FROM `tabCollections Customer Snapshot`
            WHERE collections_import = %(imp)s
        """, {"imp": doc.name})
        frappe.db.commit()

        # Find customer name column
        customer_col = None
        for raw_col, field in header_map.items():
            if field == "excel_customer_name":
                customer_col = raw_col
                break

        log_lines = []
        imported = 0
        skipped = 0

        for row in data:
            cust_name = (str(row.get(customer_col, "") or "")).strip()
            if not cust_name:
                skipped += 1
                continue

            snapshot, err = create_snapshot_row(doc, row, header_map, log_lines)
            if snapshot:
                imported += 1
            else:
                skipped += 1
                if err:
                    log_lines.append(f"  [row] {err}")

        # Update summary
        doc.reload()
        doc.sheet_name_used = sheet_name
        doc.rows_found = len(data)
        doc.rows_imported = imported
        doc.rows_skipped = skipped
        doc.status = "Imported"
        doc.validation_log = (
            f"Import complete.\n"
            f"Rows imported: {imported}\n"
            f"Rows skipped: {skipped}\n"
        )
        if log_lines:
            doc.validation_log += "\nWarnings:\n" + "\n".join(log_lines)

        doc.save(ignore_permissions=True)

        # Recalculate totals
        update_import_summary(doc)

        frappe.db.commit()

        return {
            "status": "ok",
            "message": f"Import complete. {imported} rows imported, {skipped} skipped.",
            "rows_imported": imported,
            "rows_skipped": skipped,
            "warnings": len(log_lines),
        }

    except Exception as e:
        frappe.db.rollback()
        frappe.log_error(f"Collections import error: {e}")
        try:
            doc = get_import_doc(import_name)
            doc.status = "Failed"
            doc.validation_log = str(e)
            doc.save(ignore_permissions=True)
            frappe.db.commit()
        except Exception:
            pass
        return {"status": "error", "message": str(e)}


@frappe.whitelist()
def rebuild_import_totals(import_name):
    """Recalculate the totals on the import doc from its snapshot rows."""
    try:
        doc = get_import_doc(import_name)
        update_import_summary(doc)
        return {"status": "ok", "message": "Totals recalculated."}
    except Exception as e:
        return {"status": "error", "message": str(e)}


@frappe.whitelist()
def get_import_log(import_name):
    """Return the validation log for an import."""
    try:
        doc = get_import_doc(import_name)
        return {
            "status": "ok",
            "data": {
                "name": doc.name,
                "status": doc.status,
                "validation_log": doc.validation_log,
                "rows_found": doc.rows_found,
                "rows_imported": doc.rows_imported,
                "rows_skipped": doc.rows_skipped,
                "customers_matched": doc.customers_matched,
                "customers_unmatched": doc.customers_unmatched,
                "assignment_matched": doc.assignment_matched,
                "assignment_unmatched": doc.assignment_unmatched,
                "multiple_assignments_found": doc.multiple_assignments_found,
            },
        }
    except Exception as e:
        return {"status": "error", "message": str(e)}
